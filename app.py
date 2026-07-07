"""
S9SCAN — local payment screenshot extraction
Run with:  python app.py
Everything stays on your machine. No cloud, no API keys, no database (yet).
"""

import io
import webbrowser
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import pandas as pd
from flask import Flask, abort, jsonify, request, send_file
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import UnidentifiedImageError
from pytesseract import TesseractNotFoundError
from werkzeug.utils import secure_filename

import ledger_store
from extractor import IMAGE_EXTS, archive_screenshot, process_image

ROOT = Path(__file__).parent
INBOX = ROOT / "screenshots"
ARCHIVE = ROOT / "archive"
INBOX.mkdir(exist_ok=True)
ARCHIVE.mkdir(exist_ok=True)

app = Flask(__name__, static_folder="static", static_url_path="")


# ---------- helpers ----------

def media_url(file_path: str) -> str | None:
    if not file_path:
        return None
    rel = Path(file_path).resolve().relative_to(ROOT.resolve())
    return "/media/" + quote(rel.as_posix())


def to_api_row(row: dict) -> dict:
    return {**row, "screenshot_url": media_url(row.get("file_path", ""))}


def process_and_append(paths: list[Path]) -> tuple[int, list[str]]:
    known = ledger_store.known_references()
    added = 0
    errors = []
    for p in paths:
        try:
            payment = process_image(p)
        except TesseractNotFoundError:
            errors.append(f"{p.name}: Tesseract OCR isn't installed or isn't on PATH — see README setup steps.")
            continue
        except UnidentifiedImageError:
            errors.append(f"{p.name}: doesn't look like a valid image file.")
            continue
        except Exception as e:
            errors.append(f"{p.name}: {e}")
            continue
        if (payment.utr and payment.utr in known) or (payment.transaction_id and payment.transaction_id in known):
            p.unlink(missing_ok=True)  # duplicate — skip and clean up
            continue
        archive_screenshot(payment, ARCHIVE)
        row = payment.to_row()
        row.pop("raw_text", None)
        row["processed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ledger_store.append(row)
        known.update(r for r in (payment.utr, payment.transaction_id) if r)
        added += 1
    return added, errors


def ledger_to_excel(rows: list[dict]) -> bytes:
    df = pd.DataFrame(rows)
    out = io.BytesIO()
    export = df.drop(columns=["id", "file_path"], errors="ignore").copy()
    export.insert(0, "screenshot", df.get("file_path", ""))
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Payments")
        ws = writer.sheets["Payments"]
        for row_idx, path in enumerate(df.get("file_path", []), start=2):
            cell = ws.cell(row=row_idx, column=1)
            cell.value = "open screenshot"
            cell.hyperlink = Path(path).as_uri() if path else None
            cell.font = Font(color="0E7A5F", underline="single")
        for col_idx, col in enumerate(export.columns, start=1):
            width = max(14, min(38, int(export[col].astype(str).str.len().max() or 14) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return out.getvalue()


# ---------- routes ----------

@app.get("/")
def index():
    return app.send_static_file("index.html")


@app.get("/api/payments")
def list_payments():
    return jsonify([to_api_row(r) for r in ledger_store.load_all()])


@app.get("/api/payments/<payment_id>")
def get_payment(payment_id):
    row = ledger_store.get(payment_id)
    if row is None:
        abort(404)
    return jsonify(to_api_row(row))


@app.put("/api/payments/<payment_id>")
def update_payment(payment_id):
    fields = request.get_json(force=True)
    fields = {k: v for k, v in fields.items()
              if k in ("amount", "utr", "transaction_id", "party", "date_time",
                        "direction", "app", "status")}
    fields["needs_review"] = False
    fields["review_reasons"] = ""
    if not ledger_store.update(payment_id, fields):
        abort(404)
    return jsonify(to_api_row(ledger_store.get(payment_id)))


@app.put("/api/payments/<payment_id>/notes")
def update_notes(payment_id):
    notes = (request.get_json(force=True) or {}).get("notes", "")
    if not ledger_store.update(payment_id, {"notes": notes}):
        abort(404)
    return jsonify(to_api_row(ledger_store.get(payment_id)))


@app.delete("/api/payments/<payment_id>")
def delete_payment(payment_id):
    if not ledger_store.soft_delete(payment_id):
        abort(404)
    return jsonify({"ok": True})


@app.get("/api/trash")
def trash():
    return jsonify([to_api_row(r) for r in ledger_store.list_trash()])


@app.post("/api/payments/<payment_id>/restore")
def restore_payment(payment_id):
    if not ledger_store.restore(payment_id):
        abort(404)
    return jsonify(to_api_row(ledger_store.get(payment_id)))


@app.delete("/api/payments/<payment_id>/purge")
def purge_payment(payment_id):
    row = ledger_store.purge(payment_id)
    if row is None:
        abort(404)
    file_path = row.get("file_path")
    if file_path:
        Path(file_path).unlink(missing_ok=True)
    return jsonify({"ok": True})


def _unique_dest(directory: Path, filename: str) -> Path:
    """Two uploads sharing a name would otherwise overwrite each other on disk —
    the second file to be processed would then vanish mid-batch."""
    name = secure_filename(filename) or "screenshot"
    dest = directory / name
    stem, suffix = dest.stem, dest.suffix
    i = 1
    while dest.exists():
        dest = directory / f"{stem}_{i}{suffix}"
        i += 1
    return dest


@app.post("/api/upload")
def upload():
    saved = []
    for up in request.files.getlist("files"):
        if Path(up.filename).suffix.lower() not in IMAGE_EXTS:
            continue
        dest = _unique_dest(INBOX, up.filename)
        up.save(dest)
        saved.append(dest)
    added, errors = process_and_append(saved)
    return jsonify({"added": added, "errors": errors})


@app.post("/api/process-folder")
def process_folder():
    pending = [p for p in INBOX.iterdir() if p.suffix.lower() in IMAGE_EXTS]
    added, errors = process_and_append(pending)
    return jsonify({"added": added, "errors": errors})


@app.get("/api/export")
def export():
    rows = ledger_store.load_all()
    data = ledger_to_excel(rows)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="s9scan.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/media/<path:relpath>")
def media(relpath):
    target = (ROOT / relpath).resolve()
    allowed_roots = (INBOX.resolve(), ARCHIVE.resolve())
    if not any(str(target).startswith(str(r)) for r in allowed_roots):
        abort(403)
    if not target.exists():
        abort(404)
    return send_file(target)


if __name__ == "__main__":
    PORT = 5000
    webbrowser.open(f"http://127.0.0.1:{PORT}")
    app.run(port=PORT)
